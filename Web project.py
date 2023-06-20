from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

import time

import pandas as pd
import requests
from bs4 import BeautifulSoup

import csv
import openpyxl

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

place = "Cebu"
adult = "4"
checkIn = "2023-07-18"
checkOut="2023-07-21"
starRatings="3"
minPrice="40000"
maxPrice="400000"
guestRatings = "9"

# 파파고 번역 기능

def translate_word(word):
    url = "https://openapi.naver.com/v1/papago/n2mt"
    headers = {
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "X-Naver-Client-Id": "tXDygsF4cfg_aU2rrIwe",
        "X-Naver-Client-Secret": "Q_tBdcdt9o"
    }
    data = {
        "source": "en",
        "target": "ko",
        "text": word
    }
    response = requests.post(url, headers=headers, data=data)
    
    if response.status_code == 200:
        translated_text = response.json()["message"]["result"]["translatedText"]
        return translated_text
    else:
        print("번역 실패")

def apply_translation(event):
    selected_text = event
    if selected_text:
        translated_text = translate_word(selected_text)
        print(f"번역 결과: {translated_text}")


url = f"https://hotels.naver.com/list?placeFileName=place%3A{place}&adultCnt={adult}&includeTax=false&sortField=consumerRating&sortDirection=descending&checkIn={checkIn}&checkOut={checkOut}&starRatings={starRatings}&minPrice={minPrice}&maxPrice={maxPrice}&guestRatings={guestRatings}"

driver.get(url)
driver.implicitly_wait(5)

driver.find_element(By.CSS_SELECTOR, "#__next > div > div > div > div.Contents_ListComponent__39yRH > div.Contents_result___1Z0_ > ul > li:nth-child(1) > div.Price_Price__7vul8 > a").click()
driver.implicitly_wait(5)

driver.switch_to.window(driver.window_handles[-1])

driver.find_element(By.CSS_SELECTOR, "#__next > div > div > div.common_Tab__CS5JL > div > a:nth-child(4)").click()
driver.implicitly_wait(10)

driver.close()

driver.implicitly_wait(10)

driver.switch_to.window(driver.window_handles[-1])
driver.find_element(By.CSS_SELECTOR, "#__next > div > div > div > div.Contents_ListComponent__39yRH > div.Contents_result___1Z0_ > ul > li:nth-child(2) > div.Price_Price__7vul8 > a").click()
driver.implicitly_wait(5)

before_sc = driver.execute_script("return window.scrollY")

while(True):
    driver.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.END)
    time.sleep(0.5)

    after_sc = driver.execute_script("return window.scrollY")

    if after_sc == before_sc :
        break
    else:
        before_sc = after_sc

items = driver.find_elements(By.CSS_SELECTOR, ".common_ReviewList__4cL43")

f = open(r"C:\Users\USER\Desktop\4-1\web\project.csv", 'w', encoding = 'CP949', newline = '')
csvWriter = csv.writer(f)

wb = openpyxl.Workbook()

ws = wb.create_sheet('리뷰 데이터')

ws['A1'] = '제목'
ws['B1'] = '긍정'
ws['C1'] = '부정'
row = 2

for item in items :
    name = item.find_element(By.CSS_SELECTOR, ".common_title_txt__Kn91_").text
    positive = item.find_element(By.CSS_SELECTOR, ".common_booking_txt___dAYT common_as_positive__8PlLV").text
    negative = item.find_element(By.CSS_SELECTOR, ".common_booking_txt___dAYT common_as_negative__LjbBo").text
    name = apply_translation(name)
    positive = apply_translation(positive)
    negative = apply_translation(negative)
    print(f"제목 : {name}, 긍정 : {positive}, 부정 : {negative}")
    csvWriter.writerow([name, positive, negative])

    ws[f'A{row}'] = name
    ws[f'B{row}'] = positive
    ws[f'C{row}'] = negative
    row += 1

f.close()

wb.save('./프로젝트.xlsx')
wb.close()

driver.switch_to.window(driver.window_handles[-1])
driver.find_element(By.CSS_SELECTOR,"#__next > div > div > div.common_Tab__CS5JL > div > a:nth-child(4)").click()
driver.implicitly_wait(5)

before_sc = driver.execute_script("return window.scrollY")

while(True):
    driver.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.END)
    time.sleep(0.5)

    after_sc = driver.execute_script("return window.scrollY")

    if after_sc == before_sc :
        break
    else:
        before_sc = after_sc

items = driver.find_elements(By.CSS_SELECTOR, ".common_ReviewList__4cL43")

f = open(r"C:\Users\USER\Desktop\4-1\web\project.csv", 'w', encoding = 'CP949', newline = '')
csvWriter = csv.writer(f)

wb = openpyxl.Workbook()

ws = wb.create_sheet('리뷰 데이터')

ws['D1'] = '제목'
ws['E1'] = '내용'
row = 2

for item in items :
    name = item.find_element(By.CSS_SELECTOR, ".common_title_txt__Kn91_").text
    positive = item.find_element(By.CSS_SELECTOR, ".common_booking_txt___dAYT common_as_positive__8PlLV").text
    name = apply_translation(name)
    positive = apply_translation(positive)
    print(f"제목 : {name}, 내용 : {positive}")
    csvWriter.writerow([name, positive, negative])

    ws[f'D{row}'] = name
    ws[f'E{row}'] = positive
    row += 1

f.close()

wb.save('./프로젝트.xlsx')
wb.close()

input()